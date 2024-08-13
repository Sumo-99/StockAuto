import openpyxl
import requests
import os
import traceback

from datetime import datetime

from api.models_crud import LocalFileCrud
from api.database import SessionLocal, engine, Base

# SQL Lite DB config and setup
Base.metadata.create_all(bind=engine)
db = SessionLocal()
file_handler = LocalFileCrud()


NAV_ENDPOINT = 'https://www.moneycontrol.com/mc/widget/mfnavonetimeinvestment/get_chart_value'

def get_nav(isin):
    #edge case -> empty string
    if(isin == ""):
        return "INVALID ISIN ENTERED"
    payload = {'isin': isin, 'dur': '1Y', 'classic':'true','type':'benchmark'}
    r = requests.get(NAV_ENDPOINT, params=payload)
    data = r.json()
    data = data["g1"]
    if type(data) == list:
        if len(data):
            return data[-1]['navValue']
        else: # response returns empty list for non spaced string of characters
            return "INVALID ISIN ENTERED"
    return "INVALID ISIN ENTERED" # response returns dictionary containing error msg for space separated string of characters
    

def nav_master(file_id, src_path, row_start, row_end, isin_col, nav_col, old_nav_col):
    try:
        print("Starting NAV Inference...")
        src_folder_name = os.path.dirname(src_path)
        workbook = openpyxl.load_workbook(src_path)

        sheet = workbook['monitor']
        row_end = int(row_end)
        row_start = int(row_start)
        print("Row start, end: ", row_start, row_end)
        print("Row start, end: ", type(row_start), type(row_end))  
        # # get isin val and copy old one
        for i in range(row_start,row_end+1):
            # Copy old NaVs to old NaV column
            old_nav_column = sheet.cell(row=i,column=old_nav_col)
            old_nav_column.value = sheet.cell(row=i,column=nav_col).value
            # Get New NaV from api
            isin_column = sheet.cell(row=i,column=isin_col)
            latest_nav = get_nav(isin_column.value)
            # Update NaV
            curr_nav_column = sheet.cell(row=i,column=nav_col)
            curr_nav_column.value = latest_nav

        # update dates
        print("Date update in progress...")
        old_date = sheet.cell(row=3,column=5).value
        if isinstance(old_date, str):
            old_date = datetime.strptime(old_date, "%d-%m-%Y")  
        old_date = old_date.strftime("%d-%m-%Y")
        sheet.cell(row=3, column=8).value = old_date
        new_date = datetime.today().strftime("%d-%m-%Y")
        sheet.cell(row=3,column=5).value = new_date

        # save to file
        dest_path = os.path.join(src_folder_name, f"nav_updated_file.xlsx")
        workbook.save(dest_path)
        # workbook.save(f"price_monitor_{new_date}.xlsx")

        file_status = "completed"

    except Exception as e:
        print("Exception in updating NAV ---> ", traceback.format_exc())
        file_status = "failure"
    
    finally:
        print("DB Update in progress...")
        
        # Save file status to DB
        file_handler.update_local_file(file_id, db, status = file_status)
